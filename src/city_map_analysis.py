"""
analyze_region_city_mapping.py

Standalone audit script for the scraped region/city Excel sheet
(columns: region_name, region_id, city_name, city_id).

It checks:
  1. Unique region names, and whether any region name maps to >1 region_id
  2. Unique city names, and whether any city name maps to >1 city_id
     (split into "same id repeated" vs "genuinely conflicting ids")
  3. Cities (by city_id) that appear under more than one region
  4. Rows with missing/blank region_id or city_id

Output: a single multi-sheet Excel workbook you can review manually.

Usage:
    python analyze_region_city_mapping.py <input.xlsx> [output.xlsx]

If output.xlsx is omitted, it defaults to "<input_name>_analysis.xlsx"
in the same folder as the input file.
"""

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="4472C4")
WARN_FILL = PatternFill("solid", start_color="FFC7CE")


def normalize_columns(df):
    """Map flexible/messy header names to the 4 expected canonical columns."""
    rename = {}
    for col in df.columns:
        key = str(col).strip().lower().replace(" ", "_").replace("-", "_")
        if "region" in key and "id" in key:
            rename[col] = "region_id"
        elif "region" in key:
            rename[col] = "region_name"
        elif "city" in key and "id" in key:
            rename[col] = "city_id"
        elif "city" in key:
            rename[col] = "city_name"
    df = df.rename(columns=rename)

    required = {"region_name", "region_id", "city_name", "city_id"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f"Could not find required column(s) in the input file: {missing}. "
            f"Found columns: {list(df.columns)}"
        )
    return df[["region_name", "region_id", "city_name", "city_id"]]


def write_sheet(wb, title, df, note=None, highlight_col=None):
    ws = wb.create_sheet(title)
    row = 1

    if note:
        ws.cell(row=row, column=1, value=note).font = Font(italic=True, color="555555")
        row += 1
        row += 1  # blank line after note

    if df.empty:
        ws.cell(row=row, column=1, value="(none found)").font = Font(italic=True)
        return ws

    header_row = row
    for c, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=c, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r, record in enumerate(df.itertuples(index=False), start=header_row + 1):
        for c, value in enumerate(record, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if highlight_col is not None and c == highlight_col:
                cell.fill = WARN_FILL

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(df.columns))}{header_row}"
    )

    for c, col_name in enumerate(df.columns, start=1):
        try:
            max_len = max(
                [len(str(col_name))]
                + [len(str(v)) for v in df[col_name].astype(str).tolist()]
            )
        except Exception:
            max_len = 15
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 10), 50)

    return ws


def analyze(input_path, output_path):
    df = pd.read_excel(input_path, dtype=str)
    df = normalize_columns(df)
    total_rows = len(df)

    # --- Missing / invalid rows -------------------------------------------------
    missing_mask = (
        df["region_id"].isna() | (df["region_id"].astype(str).str.strip() == "")
    ) | (df["city_id"].isna() | (df["city_id"].astype(str).str.strip() == ""))
    missing_df = df[missing_mask].copy()
    clean_df = df[~missing_mask].copy()

    # --- Unique regions + region name -> multiple region_id check --------------
    region_groups = (
        clean_df.groupby("region_name")["region_id"]
        .agg(lambda s: sorted(set(s)))
        .reset_index()
    )
    region_groups["distinct_region_id_count"] = region_groups["region_id"].apply(len)
    region_groups["region_ids"] = region_groups["region_id"].apply(
        lambda ids: ", ".join(ids)
    )
    unique_regions_df = region_groups[
        ["region_name", "region_ids", "distinct_region_id_count"]
    ].sort_values("region_name")

    dup_regions_df = unique_regions_df[
        unique_regions_df["distinct_region_id_count"] > 1
    ].sort_values("distinct_region_id_count", ascending=False)

    # --- City name -> multiple city_id check ------------------------------------
    city_groups = (
        clean_df.groupby("city_name")
        .agg(
            distinct_city_ids=("city_id", lambda s: sorted(set(s))),
            row_count=("city_id", "count"),
            regions_seen=("region_name", lambda s: sorted(set(s))),
        )
        .reset_index()
    )
    city_groups["distinct_city_id_count"] = city_groups["distinct_city_ids"].apply(len)
    city_groups["city_ids"] = city_groups["distinct_city_ids"].apply(
        lambda ids: ", ".join(ids)
    )
    city_groups["regions"] = city_groups["regions_seen"].apply(lambda r: ", ".join(r))

    repeated_cities = city_groups[city_groups["row_count"] > 1]

    dup_cities_same_id_df = repeated_cities[
        repeated_cities["distinct_city_id_count"] == 1
    ][["city_name", "city_ids", "row_count", "regions"]].sort_values(
        "row_count", ascending=False
    )

    dup_cities_diff_id_df = repeated_cities[
        repeated_cities["distinct_city_id_count"] > 1
    ][
        ["city_name", "city_ids", "distinct_city_id_count", "row_count", "regions"]
    ].sort_values("distinct_city_id_count", ascending=False)

    # --- Same city_id appearing under multiple regions --------------------------
    id_region_groups = (
        clean_df.groupby("city_id")
        .agg(
            city_names=("city_name", lambda s: sorted(set(s))),
            distinct_regions=("region_name", lambda s: sorted(set(s))),
        )
        .reset_index()
    )
    id_region_groups["distinct_region_count"] = id_region_groups[
        "distinct_regions"
    ].apply(len)
    multi_region_cities_df = id_region_groups[
        id_region_groups["distinct_region_count"] > 1
    ].copy()
    multi_region_cities_df["city_names"] = multi_region_cities_df["city_names"].apply(
        lambda x: ", ".join(x)
    )
    multi_region_cities_df["regions"] = multi_region_cities_df[
        "distinct_regions"
    ].apply(lambda x: ", ".join(x))
    multi_region_cities_df = multi_region_cities_df[
        ["city_id", "city_names", "distinct_region_count", "regions"]
    ].sort_values("distinct_region_count", ascending=False)

    # --- Summary ------------------------------------------------------------------
    summary_rows = [
        ("Total rows in input file", total_rows),
        ("Rows with missing region_id or city_id", len(missing_df)),
        ("Rows used in analysis (after excluding missing)", len(clean_df)),
        ("", ""),
        ("Unique region names", unique_regions_df.shape[0]),
        ("Region names mapped to >1 region_id (conflicts)", dup_regions_df.shape[0]),
        ("", ""),
        ("Unique city names", city_groups.shape[0]),
        (
            "City names that repeat with the SAME city_id",
            dup_cities_same_id_df.shape[0],
        ),
        (
            "City names that repeat with DIFFERENT city_ids (conflicts)",
            dup_cities_diff_id_df.shape[0],
        ),
        (
            "city_ids that appear under more than one region",
            multi_region_cities_df.shape[0],
        ),
    ]
    summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    # --- Build workbook -------------------------------------------------------
    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(wb, "Summary", summary_df)

    write_sheet(
        wb,
        "Unique Regions",
        unique_regions_df,
        note="One row per distinct region_name. distinct_region_id_count > 1 means that name maps to more than one region_id.",
        highlight_col=3,
    )

    write_sheet(
        wb,
        "Duplicate Regions",
        dup_regions_df,
        note="Region names associated with more than one region_id. Likely scrape/data inconsistencies worth resolving manually.",
    )

    write_sheet(
        wb,
        "Dup Cities - Same ID",
        dup_cities_same_id_df,
        note="City names that appear multiple times in the sheet, but consistently with the SAME city_id (likely harmless duplicate rows).",
    )

    write_sheet(
        wb,
        "Dup Cities - Diff ID",
        dup_cities_diff_id_df,
        note="City names that appear with MORE THAN ONE distinct city_id. These are genuine conflicts and should be reviewed manually before mapping.",
    )

    write_sheet(
        wb,
        "Cities Multi-Region",
        multi_region_cities_df,
        note="Same city_id appearing under more than one region_name. May indicate scrape/data quality issues (e.g. mis-categorized regions).",
    )

    write_sheet(
        wb,
        "Missing or Invalid Rows",
        missing_df,
        note="Rows excluded from the analysis above because region_id or city_id was blank.",
    )

    wb.save(output_path)
    return output_path, summary_df


def main():
    if len(sys.argv) < 2:
        print("Usage: python analyze_region_city_mapping.py <input.xlsx> [output.xlsx]")
        sys.exit(1)

    input_path = sys.argv[1]
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        base, _ = os.path.splitext(input_path)
        output_path = f"{base}_analysis.xlsx"

    output_path, summary_df = analyze(input_path, output_path)

    print(f"Analysis written to: {output_path}\n")
    print(summary_df.to_string(index=False))


if __name__ == "__main__":
    main()
